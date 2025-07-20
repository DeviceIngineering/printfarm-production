"""
Excel exporters for PrintFarm production system.
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

class ProductsExporter:
    """
    Export products list to Excel with PrintFarm branding.
    """
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Товары"
        
        # PrintFarm brand colors
        self.brand_color = "06EAFC"  # PrintFarm cyan
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color=self.brand_color, end_color=self.brand_color, fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.title_font = Font(bold=True, size=16, color="1E1E1E")
        self.subtitle_font = Font(size=12, italic=True, color="595959")
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Type colors
        self.type_fills = {
            'new': PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
            'old': PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
            'critical': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
        }
        
        # Number format
        self.number_format = '#,##0.00'
    
    def export_products(self, products_queryset) -> HttpResponse:
        """
        Export products to Excel file with formatting.
        """
        # Add title
        self.ws.merge_cells('A1:K1')
        self.ws['A1'] = 'PRINTFARM - СПИСОК ТОВАРОВ'
        self.ws['A1'].font = self.title_font
        self.ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[1].height = 30
        
        # Convert to list to handle both querysets and sliced querysets
        products_list = list(products_queryset)
        
        # Add metadata
        self.ws['A3'] = f'Дата выгрузки: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
        self.ws['A3'].font = self.subtitle_font
        self.ws['A4'] = f'Всего товаров: {len(products_list)}'
        self.ws['A4'].font = self.subtitle_font
        
        # Calculate statistics
        total_stock = sum(float(p.current_stock) for p in products_list)
        need_production = sum(1 for p in products_list if p.production_needed > 0)
        critical_count = sum(1 for p in products_list if p.product_type == 'critical')
        
        self.ws['A5'] = f'Общий остаток: {total_stock:,.0f} ед.'
        self.ws['A5'].font = self.subtitle_font
        self.ws['A6'] = f'Требуют производства: {need_production} поз.'
        self.ws['A6'].font = self.subtitle_font
        self.ws['A7'] = f'Критических позиций: {critical_count}'
        self.ws['A7'].font = Font(bold=True, color="FF0055")
        
        # Headers
        headers = [
            ('Артикул', 15),
            ('Название', 40),
            ('Тип', 12),
            ('Группа', 25),
            ('Текущий\nостаток', 12),
            ('Продажи\nза 2 мес.', 12),
            ('Средн.\nпотр./день', 12),
            ('Дней\nостатка', 10),
            ('Нужно\nпроизвести', 12),
            ('Приоритет', 10),
            ('Последняя\nсинхронизация', 18)
        ]
        
        row_start = 9
        for col, (header, width) in enumerate(headers, 1):
            cell = self.ws.cell(row=row_start, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
            self.ws.column_dimensions[get_column_letter(col)].width = width
        
        self.ws.row_dimensions[row_start].height = 40
        
        # Data
        for row_num, product in enumerate(products_list, row_start + 1):
            # Article
            self.ws.cell(row=row_num, column=1, value=product.article).border = self.border
            
            # Name
            self.ws.cell(row=row_num, column=2, value=product.name).border = self.border
            
            # Type
            type_cell = self.ws.cell(row=row_num, column=3, value=self._get_product_type_display(product.product_type))
            type_cell.border = self.border
            if product.product_type in self.type_fills:
                type_cell.fill = self.type_fills[product.product_type]
            
            # Group
            self.ws.cell(row=row_num, column=4, value=product.product_group_name or '-').border = self.border
            
            # Current stock
            stock_cell = self.ws.cell(row=row_num, column=5, value=float(product.current_stock))
            stock_cell.border = self.border
            stock_cell.number_format = self.number_format
            stock_cell.alignment = Alignment(horizontal="right")
            
            # Sales last 2 months
            sales_cell = self.ws.cell(row=row_num, column=6, value=float(product.sales_last_2_months))
            sales_cell.border = self.border
            sales_cell.number_format = self.number_format
            sales_cell.alignment = Alignment(horizontal="right")
            
            # Average daily consumption
            consumption_cell = self.ws.cell(row=row_num, column=7, value=float(product.average_daily_consumption))
            consumption_cell.border = self.border
            consumption_cell.number_format = '#,##0.0000'
            consumption_cell.alignment = Alignment(horizontal="right")
            
            # Days of stock
            days_cell = self.ws.cell(row=row_num, column=8)
            if product.days_of_stock is not None:
                days_cell.value = float(product.days_of_stock)
                days_cell.number_format = '#,##0.0'
                if product.days_of_stock < 5:
                    days_cell.font = Font(bold=True, color="FF0055")
                elif product.days_of_stock < 10:
                    days_cell.font = Font(color="FFB800")
            else:
                days_cell.value = '-'
            days_cell.border = self.border
            days_cell.alignment = Alignment(horizontal="center")
            
            # Production needed
            prod_cell = self.ws.cell(row=row_num, column=9, value=float(product.production_needed))
            prod_cell.border = self.border
            prod_cell.number_format = self.number_format
            prod_cell.alignment = Alignment(horizontal="right")
            if product.production_needed > 0:
                prod_cell.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
            
            # Priority
            priority_cell = self.ws.cell(row=row_num, column=10, value=product.production_priority)
            priority_cell.border = self.border
            priority_cell.alignment = Alignment(horizontal="center")
            if product.production_priority >= 80:
                priority_cell.font = Font(bold=True, color="FF0055")
            elif product.production_priority >= 60:
                priority_cell.font = Font(color="FFB800")
            
            # Last synced
            sync_cell = self.ws.cell(row=row_num, column=11)
            if product.last_synced_at:
                sync_cell.value = product.last_synced_at.strftime("%d.%m.%Y %H:%M")
            else:
                sync_cell.value = '-'
            sync_cell.border = self.border
            sync_cell.alignment = Alignment(horizontal="center")
        
        # Add legend
        legend_row = row_start + len(products_list) + 3
        self.ws.merge_cells(f'A{legend_row}:D{legend_row}')
        self.ws[f'A{legend_row}'] = 'Легенда:'
        self.ws[f'A{legend_row}'].font = Font(bold=True)
        
        legend_items = [
            ('Новая позиция', self.type_fills['new']),
            ('Старая позиция', self.type_fills['old']),
            ('Критическая позиция', self.type_fills['critical']),
        ]
        
        for i, (text, fill) in enumerate(legend_items):
            row = legend_row + i + 1
            cell = self.ws[f'B{row}']
            cell.value = text
            cell.fill = fill
            cell.border = self.border
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'printfarm_products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        self.wb.save(response)
        return response
    
    def _get_product_type_display(self, product_type: str) -> str:
        """
        Get display name for product type.
        """
        type_map = {
            'new': 'Новая',
            'old': 'Старая',
            'critical': 'Критическая'
        }
        return type_map.get(product_type, product_type)
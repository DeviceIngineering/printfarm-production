from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from django.db.models import Q
from apps.products.models import Product
from apps.products.serializers import ProductListSerializer
from apps.products.services.reserve_calculator import ReserveCalculatorService
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from datetime import datetime

def normalize_article(article):
    """
    Нормализация артикула для корректного сравнения
    - Удаляет лишние пробелы в начале и конце
    - Заменяет все виды тире на обычное тире
    - Удаляет невидимые символы
    - Приводит к стандартному виду
    """
    if not article:
        return ''
    
    # Конвертируем в строку и убираем пробелы
    article = str(article).strip()
    
    # Заменяем различные виды тире и дефисов на обычный дефис
    # Включает: en dash (–), em dash (—), minus (−), hyphen-minus (-)
    article = re.sub(r'[–—−‒]', '-', article)
    
    # Удаляем невидимые символы (кроме обычного пробела)
    article = re.sub(r'[\x00-\x1f\x7f-\x9f\xa0\u2000-\u200f\u2028-\u202f\u205f-\u206f]', '', article)
    
    # Удаляем лишние пробелы внутри строки
    article = re.sub(r'\s+', ' ', article).strip()
    
    return article

@api_view(['GET'])
@permission_classes([AllowAny])
def get_products_for_tochka(request):
    """
    API для получения списка всех товаров для вкладки Точка
    Поддерживает параметр include_reserve для отображения колонки Резерв
    """
    try:
        # Получаем все товары с базовой информацией
        products = Product.objects.all().order_by('-updated_at')
        
        # Применяем пагинацию если нужно
        limit = request.GET.get('limit')
        if limit:
            try:
                limit = int(limit)
                products = products[:limit]
            except ValueError:
                pass
        
        # Получаем параметр include_reserve
        include_reserve = request.GET.get('include_reserve', '').lower() in ['true', '1', 'yes']
        
        # Сериализация данных с контекстом include_reserve
        serializer = ProductListSerializer(
            products, 
            many=True, 
            context={
                'request': request,
                'include_reserve': include_reserve
            }
        )
        
        return Response({
            'results': serializer.data,
            'count': len(serializer.data),
            'include_reserve': include_reserve,
            'message': 'Товары успешно загружены'
        })
        
    except Exception as e:
        return Response({
            'error': f'Ошибка при загрузке товаров: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_production_list_for_tochka(request):
    """
    API для получения списка товаров на производство для вкладки Точка
    Поддерживает параметр include_reserve для отображения колонки Резерв
    """
    try:
        # Получаем товары, которые требуют производства ИЛИ имеют резерв
        # Бизнес-правило: товары с резервом всегда включаются в планирование производства
        from django.db.models import Q
        products = Product.objects.filter(
            Q(production_needed__gt=0) | Q(reserved_stock__gt=0)
        ).order_by('-production_priority', 'article')
        
        # Применяем пагинацию если нужно
        limit = request.GET.get('limit')
        if limit:
            try:
                limit = int(limit)
                products = products[:limit]
            except ValueError:
                pass
        
        # Получаем параметр include_reserve
        include_reserve = request.GET.get('include_reserve', '').lower() in ['true', '1', 'yes']
        
        # Сериализация данных с контекстом include_reserve
        serializer = ProductListSerializer(
            products, 
            many=True, 
            context={
                'request': request,
                'include_reserve': include_reserve
            }
        )
        
        # Добавляем информацию о товарах с резервом для подсветки
        results_with_highlight = []
        for item in serializer.data:
            # Проверяем есть ли резерв у товара
            reserved_stock = float(item.get('reserved_stock', 0))
            has_reserve = reserved_stock > 0
            
            # Добавляем флаг для подсветки в UI
            item_with_highlight = dict(item)
            item_with_highlight['has_reserve'] = has_reserve
            item_with_highlight['reserve_amount'] = reserved_stock
            
            # Используем новый алгоритм расчета резерва с цветовой индикацией
            if include_reserve:
                current_stock = float(item.get('current_stock', 0))
                reserve_calculator = ReserveCalculatorService()
                reserve_calc = reserve_calculator.calculate_reserve_display(
                    reserved_stock=reserved_stock,
                    current_stock=current_stock
                )
                reserve_ui = reserve_calculator.format_reserve_for_display(reserve_calc)
                
                # Добавляем результаты расчета резерва
                item_with_highlight['calculated_reserve'] = float(reserve_calc['calculated_reserve'])
                item_with_highlight['reserve_color'] = reserve_calc['color_indicator']
                item_with_highlight['reserve_display_text'] = reserve_ui['display_text']
                item_with_highlight['reserve_tooltip'] = reserve_ui['tooltip_text']
                item_with_highlight['reserve_needs_attention'] = reserve_ui['needs_attention']
                
                # Обратная совместимость со старым полем
                item_with_highlight['reserve_minus_stock'] = float(reserve_calc['calculated_reserve'])
            else:
                item_with_highlight['calculated_reserve'] = None
                item_with_highlight['reserve_color'] = 'gray'
                item_with_highlight['reserve_display_text'] = None
                item_with_highlight['reserve_tooltip'] = None
                item_with_highlight['reserve_needs_attention'] = False
                item_with_highlight['reserve_minus_stock'] = None
            
            results_with_highlight.append(item_with_highlight)
        
        return Response({
            'results': results_with_highlight,
            'count': len(results_with_highlight),
            'include_reserve': include_reserve,
            'message': 'Список на производство успешно загружен'
        })
        
    except Exception as e:
        return Response({
            'error': f'Ошибка при загрузке списка на производство: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_products_stats_for_tochka(request):
    """
    API для получения статистики товаров для вкладки Точка
    """
    try:
        total_products = Product.objects.count()
        production_needed = Product.objects.filter(production_needed__gt=0).count()
        critical_products = Product.objects.filter(product_type='critical').count()
        new_products = Product.objects.filter(product_type='new').count()
        old_products = Product.objects.filter(product_type='old').count()
        
        return Response({
            'total_products': total_products,
            'production_needed': production_needed,
            'critical_products': critical_products,
            'new_products': new_products,
            'old_products': old_products,
            'message': 'Статистика успешно загружена'
        })
        
    except Exception as e:
        return Response({
            'error': f'Ошибка при загрузке статистики: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def upload_excel_file_for_tochka(request):
    """
    API для загрузки Excel файла с данными для вкладки Точка
    Ищет колонки "Артикул товара" и "Заказов, шт."
    """
    try:
        if 'file' not in request.FILES:
            return Response({
                'error': 'Файл не найден. Пожалуйста, выберите Excel файл.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Проверяем расширение файла
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({
                'error': 'Неверный формат файла. Поддерживаются только .xlsx и .xls файлы.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Читаем Excel файл
        try:
            if excel_file.name.endswith('.xlsx'):
                df = pd.read_excel(io.BytesIO(excel_file.read()), engine='openpyxl')
            else:
                df = pd.read_excel(io.BytesIO(excel_file.read()), engine='xlrd')
        except Exception as e:
            return Response({
                'error': f'Ошибка при чтении Excel файла: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Ищем нужные колонки
        article_column = None
        orders_column = None
        
        # Возможные варианты названий колонок
        article_variants = ['Артикул товара', 'артикул товара', 'Артикул', 'артикул', 'Article', 'article']
        orders_variants = ['Заказов, шт.', 'заказов, шт.', 'Заказов шт', 'заказов шт', 'Заказов', 'заказов', 'Orders', 'orders']
        
        # Поиск колонки с артикулами
        for col in df.columns:
            if str(col).strip() in article_variants:
                article_column = col
                break
        
        # Поиск колонки с заказами
        for col in df.columns:
            if str(col).strip() in orders_variants:
                orders_column = col
                break
        
        if article_column is None:
            return Response({
                'error': 'Колонка "Артикул товара" не найдена в файле.',
                'available_columns': list(df.columns)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if orders_column is None:
            return Response({
                'error': 'Колонка "Заказов, шт." не найдена в файле.',
                'available_columns': list(df.columns)
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Извлекаем данные из нужных колонок
        raw_data = []
        processed_count = 0
        error_count = 0
        
        for index, row in df.iterrows():
            try:
                # Используем функцию нормализации для артикула
                raw_article = row[article_column] if pd.notna(row[article_column]) else ''
                article = normalize_article(raw_article)
                orders = row[orders_column] if pd.notna(row[orders_column]) else 0
                
                # Пропускаем пустые артикулы
                if not article or article.lower() in ['nan', 'none', '']:
                    continue
                
                # Пытаемся преобразовать количество заказов в число
                try:
                    orders = float(orders)
                    if orders < 0:
                        orders = 0
                except (ValueError, TypeError):
                    orders = 0
                
                raw_data.append({
                    'article': article,
                    'orders': int(orders),
                    'row_number': index + 1
                })
                processed_count += 1
                
            except Exception as e:
                error_count += 1
                continue
        
        if not raw_data:
            return Response({
                'error': 'Не удалось извлечь данные из файла. Проверьте формат данных.',
                'processed_count': processed_count,
                'error_count': error_count
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Дедупликация по артикулу с суммированием заказов
        article_dict = {}
        duplicate_count = 0
        
        for item in raw_data:
            article = item['article']
            orders = item['orders']
            
            if article in article_dict:
                # Суммируем заказы для дублирующихся артикулов
                article_dict[article]['orders'] += orders
                article_dict[article]['duplicate_rows'].append(item['row_number'])
                duplicate_count += 1
            else:
                # Первое вхождение артикула
                article_dict[article] = {
                    'article': article,
                    'orders': orders,
                    'row_number': item['row_number'],
                    'duplicate_rows': []
                }
        
        # Формируем финальный список без дубликатов
        extracted_data = []
        for article, data in article_dict.items():
            extracted_data.append({
                'article': data['article'],
                'orders': data['orders'],
                'row_number': data['row_number'],
                'has_duplicates': len(data['duplicate_rows']) > 0,
                'duplicate_rows': data['duplicate_rows'] if data['duplicate_rows'] else None
            })
        
        # Сортируем по убыванию количества заказов
        extracted_data.sort(key=lambda x: x['orders'], reverse=True)
        
        return Response({
            'message': f'Файл успешно обработан. Уникальных артикулов: {len(extracted_data)}, дубликатов обработано: {duplicate_count}.',
            'data': extracted_data,  # Показываем все записи без ограничений
            'total_records': len(extracted_data),
            'total_raw_records': len(raw_data),
            'unique_articles': len(extracted_data),
            'duplicates_merged': duplicate_count,
            'processed_count': processed_count,
            'error_count': error_count,
            'columns_found': {
                'article_column': article_column,
                'orders_column': orders_column
            }
        })
        
    except Exception as e:
        return Response({
            'error': f'Внутренняя ошибка сервера: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def merge_excel_with_products(request):
    """
    API для объединения данных из Excel с товарами по артикулу
    Excel содержит товары Точки, нужно найти товары МойСклад которых НЕТ в Точке
    """
    try:
        excel_data = request.data.get('excel_data', [])
        
        if not excel_data:
            return Response({
                'error': 'Нет данных Excel для объединения'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Нормализуем артикулы из Excel данных и создаем словари
        excel_articles = set()
        excel_dict = {}
        normalized_to_original = {}  # Словарь для обратного поиска
        
        for item in excel_data:
            if item.get('article'):
                original_article = item['article']
                normalized_article = normalize_article(original_article)
                
                if normalized_article:  # Если артикул не пустой после нормализации
                    excel_articles.add(normalized_article)
                    excel_dict[normalized_article] = item
                    normalized_to_original[normalized_article] = original_article
        
        if not excel_articles:
            return Response({
                'error': 'Не найдены артикулы в данных Excel'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        
        # Загружаем ВСЕ товары из базы данных (МойСклад) которые требуют производства (включая товары с резервом)
        from django.db.models import Q
        products_for_production = Product.objects.filter(
            Q(production_needed__gt=0) | Q(reserved_stock__gt=0)
        ).order_by('-production_priority', 'article')
        
        
        # Объединяем данные
        merged_data = []
        in_tochka_count = 0
        not_in_tochka_count = 0
        
        for product in products_for_production:
            # Нормализуем артикул из базы данных для сравнения
            normalized_product_article = normalize_article(product.article)
            
            # Проверяем есть ли товар в Excel (Точке)
            excel_item = excel_dict.get(normalized_product_article)
            
            
            if excel_item:
                # Товар ЕСТЬ в Точке - объединяем данные
                merged_item = {
                    # Данные из товаров МойСклад
                    'article': product.article,
                    'product_id': product.id,
                    'product_name': product.name,
                    'product_description': product.description,
                    'current_stock': float(product.current_stock),
                    'sales_last_2_months': float(product.sales_last_2_months),
                    'product_type': product.product_type,
                    'production_needed': float(product.production_needed),
                    'production_priority': product.production_priority,
                    'days_of_stock': float(product.days_of_stock) if product.days_of_stock else None,
                    
                    # Данные из Excel (Точка)
                    'orders_in_tochka': excel_item.get('orders', 0),
                    'excel_row_number': excel_item.get('row_number', 0),
                    'has_duplicates': excel_item.get('has_duplicates', False),
                    'duplicate_rows': excel_item.get('duplicate_rows'),
                    
                    # Статус
                    'is_in_tochka': True,
                    'needs_registration': False,
                }
                in_tochka_count += 1
            else:
                # Товара НЕТ в Точке - помечаем специально
                merged_item = {
                    # Данные из товаров МойСклад
                    'article': product.article,
                    'product_id': product.id,
                    'product_name': product.name,
                    'product_description': product.description,
                    'current_stock': float(product.current_stock),
                    'sales_last_2_months': float(product.sales_last_2_months),
                    'product_type': product.product_type,
                    'production_needed': float(product.production_needed),
                    'production_priority': product.production_priority,
                    'days_of_stock': float(product.days_of_stock) if product.days_of_stock else None,
                    
                    # Пустые данные Точки
                    'orders_in_tochka': 0,
                    'excel_row_number': None,
                    'has_duplicates': False,
                    'duplicate_rows': None,
                    
                    # Статус
                    'is_in_tochka': False,
                    'needs_registration': True,  # Нужно завести в Точке!
                }
                not_in_tochka_count += 1
            
            merged_data.append(merged_item)
        
        # Сортируем: сначала товары которых нет в Точке, потом по приоритету производства
        merged_data.sort(key=lambda x: (not x['needs_registration'], -x['production_priority']))
        
        return Response({
            'message': f'Анализ завершен. Товаров в Точке: {in_tochka_count}, отсутствует в Точке: {not_in_tochka_count}',
            'data': merged_data,
            'total_production_needed': len(merged_data),
            'products_in_tochka': in_tochka_count,
            'products_not_in_tochka': not_in_tochka_count,
            'coverage_rate': round((in_tochka_count / len(merged_data)) * 100, 1) if merged_data else 0,
        })
        
    except Exception as e:
        return Response({
            'error': f'Ошибка при объединении данных: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def get_filtered_production_list(request):
    """
    API для получения отфильтрованного списка на производство
    Возвращает только товары которые ЕСТЬ в Точке (исключает товары отсутствующие в Точке)
    Поддерживает параметр include_reserve для отображения колонки Резерв
    """
    try:
        excel_data = request.data.get('excel_data', [])
        include_reserve = request.data.get('include_reserve', False)
        
        if not excel_data:
            return Response({
                'error': 'Нет данных Excel для фильтрации'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Нормализуем артикулы из Excel данных
        excel_articles = set()
        excel_dict = {}
        
        for item in excel_data:
            if item.get('article'):
                normalized_article = normalize_article(item['article'])
                if normalized_article:
                    excel_articles.add(normalized_article)
                    excel_dict[normalized_article] = item
        
        if not excel_articles:
            return Response({
                'error': 'Не найдены артикулы в данных Excel'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Загружаем ВСЕ товары на производство (включая товары с резервом) и фильтруем те, которые ЕСТЬ в Точке
        from django.db.models import Q
        all_products_for_production = Product.objects.filter(
            Q(production_needed__gt=0) | Q(reserved_stock__gt=0)
        ).order_by('-production_priority', 'article')
        
        # Фильтруем товары, которые есть в Excel (Точке)
        products_in_tochka = []
        for product in all_products_for_production:
            normalized_product_article = normalize_article(product.article)
            if normalized_product_article in excel_articles:
                products_in_tochka.append(product)
        
        # Используем сериализатор для корректной обработки резерва
        serializer = ProductListSerializer(
            products_in_tochka, 
            many=True, 
            context={
                'request': request,
                'include_reserve': include_reserve
            }
        )
        
        # Обогащаем данные Excel информацией и подсветкой
        production_list = []
        total_quantity = 0
        
        for product_data in serializer.data:
            normalized_product_article = normalize_article(product_data['article'])
            excel_item = excel_dict.get(normalized_product_article, {})
            
            # Добавляем Excel данные
            item = dict(product_data)
            item.update({
                'orders_in_tochka': excel_item.get('orders', 0),
                'has_duplicates': excel_item.get('has_duplicates', False),
            })
            
            # Добавляем информацию для подсветки
            reserved_stock = float(item.get('reserved_stock', 0))
            has_reserve = reserved_stock > 0
            item['has_reserve'] = has_reserve
            item['reserve_amount'] = reserved_stock
            
            # Используем новый алгоритм расчета резерва с цветовой индикацией
            if include_reserve:
                current_stock = float(item.get('current_stock', 0))
                reserve_calculator = ReserveCalculatorService()
                reserve_calc = reserve_calculator.calculate_reserve_display(
                    reserved_stock=reserved_stock,
                    current_stock=current_stock
                )
                reserve_ui = reserve_calculator.format_reserve_for_display(reserve_calc)
                
                # Добавляем результаты расчета резерва
                item['calculated_reserve'] = float(reserve_calc['calculated_reserve'])
                item['reserve_color'] = reserve_calc['color_indicator']
                item['reserve_display_text'] = reserve_ui['display_text']
                item['reserve_tooltip'] = reserve_ui['tooltip_text']
                item['reserve_needs_attention'] = reserve_ui['needs_attention']
                
                # Обратная совместимость со старым полем
                item['reserve_minus_stock'] = float(reserve_calc['calculated_reserve'])
            else:
                item['calculated_reserve'] = None
                item['reserve_color'] = 'gray'
                item['reserve_display_text'] = None
                item['reserve_tooltip'] = None
                item['reserve_needs_attention'] = False
                item['reserve_minus_stock'] = None
            
            production_list.append(item)
            total_quantity += float(item.get('production_needed', 0))
        
        return Response({
            'message': f'Отфильтрованный список готов: {len(production_list)} товаров к производству',
            'data': production_list,
            'total_items': len(production_list),
            'total_quantity': round(total_quantity, 2),
            'include_reserve': include_reserve,
            'filtered_by_tochka': True,
        })
        
    except Exception as e:
        return Response({
            'error': f'Ошибка при создании отфильтрованного списка: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def export_deduplicated_excel(request):
    """
    API для экспорта дедуплицированных данных Excel в файл
    """
    try:
        data = request.data.get('data', [])
        
        if not data:
            return Response({
                'error': 'Нет данных для экспорта'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Данные Excel без дублей"
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="06EAFC", end_color="06EAFC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Заголовки колонок
        headers = ['Артикул', 'Заказов (шт.)', 'Номер строки', 'Есть дубликаты', 'Строки дубликатов']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные
        for row, item in enumerate(data, 2):
            ws.cell(row=row, column=1, value=item.get('article', ''))
            ws.cell(row=row, column=2, value=item.get('orders', 0))
            ws.cell(row=row, column=3, value=item.get('row_number', ''))
            ws.cell(row=row, column=4, value='Да' if item.get('has_duplicates') else 'Нет')
            
            # Объединяем номера строк дубликатов
            duplicate_rows = item.get('duplicate_rows', [])
            if duplicate_rows:
                ws.cell(row=row, column=5, value=', '.join(map(str, duplicate_rows)))
            else:
                ws.cell(row=row, column=5, value='')
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Подготавливаем ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'Данные_Excel_без_дублей_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return Response({
            'error': f'Ошибка при экспорте: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def export_production_list(request):
    """
    API для экспорта списка к производству в Excel файл
    """
    try:
        data = request.data.get('data', [])
        
        if not data:
            return Response({
                'error': 'Нет данных для экспорта'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Список к производству"
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="13C2C2", end_color="13C2C2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Заголовки колонок
        headers = [
            'Артикул', 'Название товара', 'К производству (шт.)', 'Приоритет', 
            'Текущий остаток', 'Тип товара', 'Продажи за 2 мес.', 'Заказов в Точке'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные
        total_production = 0
        for row, item in enumerate(data, 2):
            ws.cell(row=row, column=1, value=item.get('article', ''))
            ws.cell(row=row, column=2, value=item.get('product_name', ''))
            
            production_needed = item.get('production_needed', 0)
            ws.cell(row=row, column=3, value=production_needed)
            total_production += production_needed
            
            ws.cell(row=row, column=4, value=item.get('production_priority', 0))
            ws.cell(row=row, column=5, value=item.get('current_stock', 0))
            
            # Переводим тип товара на русский
            product_type = item.get('product_type', '')
            type_translation = {
                'new': 'Новый',
                'old': 'Старый', 
                'critical': 'Критичный'
            }
            ws.cell(row=row, column=6, value=type_translation.get(product_type, product_type))
            
            ws.cell(row=row, column=7, value=item.get('sales_last_2_months', 0))
            ws.cell(row=row, column=8, value=item.get('orders_in_tochka', 0))
        
        # Добавляем итоговую строку
        total_row = len(data) + 3
        ws.cell(row=total_row, column=2, value="ИТОГО К ПРОИЗВОДСТВУ:")
        ws.cell(row=total_row, column=3, value=total_production)
        
        # Стиль для итоговой строки
        for col in [2, 3]:
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Подготавливаем ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'Список_к_производству_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return Response({
            'error': f'Ошибка при экспорте: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




@api_view(['POST'])
@permission_classes([AllowAny])
def upload_and_auto_process_excel(request):
    """
    API для автоматической обработки Excel файла:
    1. Загружает и дедуплицирует Excel
    2. Автоматически выполняет анализ производства
    3. Автоматически формирует список к производству
    
    Возвращает полный результат всех операций
    """
    import time
    start_time = time.time()
    
    try:
        # Проверяем, что файл передан
        if 'file' not in request.FILES:
            return Response({
                'error': 'Файл не был загружен'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Этап 1: Загрузка и дедупликация Excel файла
        try:
            file_content = excel_file.read()
            df = pd.read_excel(io.BytesIO(file_content))
            
            # Поиск нужных колонок с различными вариантами названий
            article_column = None
            orders_column = None
            
            for col in df.columns:
                col_lower = str(col).lower().strip()
                if 'артикул' in col_lower and 'товар' in col_lower:
                    article_column = col
                elif 'заказ' in col_lower and 'шт' in col_lower:
                    orders_column = col
            
            if not article_column or not orders_column:
                return Response({
                    'error': 'Не найдены необходимые колонки "Артикул товара" и "Заказов, шт."'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Фильтруем и очищаем данные
            df_filtered = df[[article_column, orders_column]].copy()
            df_filtered = df_filtered.dropna()
            
            # Приводим к нужным типам
            df_filtered[article_column] = df_filtered[article_column].astype(str)
            df_filtered[orders_column] = pd.to_numeric(df_filtered[orders_column], errors='coerce')
            df_filtered = df_filtered.dropna()
            
            # Дедупликация по артикулу с суммированием заказов
            deduplicated_data = []
            article_groups = df_filtered.groupby(article_column)
            
            for article, group in article_groups:
                normalized_article = normalize_article(article)
                if not normalized_article:
                    continue
                    
                total_orders = group[orders_column].sum()
                row_numbers = group.index.tolist()
                
                item = {
                    'article': normalized_article,
                    'orders': int(total_orders),
                    'row_number': row_numbers[0] + 2,  # +2 для Excel нумерации (1-based + header)
                    'has_duplicates': len(row_numbers) > 1,
                    'duplicate_rows': row_numbers[1:] if len(row_numbers) > 1 else []
                }
                deduplicated_data.append(item)
            
            # Сортируем по убыванию количества заказов
            deduplicated_data.sort(key=lambda x: x['orders'], reverse=True)
            
            upload_result = {
                'message': f'Excel файл обработан успешно',
                'total_records': len(df_filtered),
                'unique_articles': len(deduplicated_data),
                'data': deduplicated_data
            }
            
        except Exception as e:
            return Response({
                'error': f'Ошибка при обработке Excel файла: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Этап 2: Автоматический анализ производства
        try:
            excel_articles = [item['article'] for item in deduplicated_data]
            
            # Получаем товары из базы данных
            products = Product.objects.filter(article__in=excel_articles)
            products_dict = {}
            
            for product in products:
                normalized_article = normalize_article(product.article)
                if normalized_article:
                    products_dict[normalized_article] = product
            
            # Объединяем данные Excel с товарами
            merged_data = []
            found_products = 0
            
            for excel_item in deduplicated_data:
                product = products_dict.get(excel_item['article'])
                
                if product:
                    found_products += 1
                    merged_item = {
                        'article': excel_item['article'],
                        'orders': excel_item['orders'],
                        'orders_in_tochka': excel_item['orders'],
                        'has_duplicates': excel_item.get('has_duplicates', False),
                        'product_name': product.name,
                        'current_stock': float(product.current_stock),
                        'sales_last_2_months': float(product.sales_last_2_months),
                        'product_type': product.product_type,
                        'production_needed': float(product.production_needed),
                        'production_priority': product.production_priority,
                        'product_matched': True,
                        'has_product_data': True,
                        'is_in_tochka': True,
                        'needs_registration': False,
                    }
                else:
                    merged_item = {
                        'article': excel_item['article'],
                        'orders': excel_item['orders'],
                        'orders_in_tochka': excel_item['orders'],
                        'has_duplicates': excel_item.get('has_duplicates', False),
                        'product_name': None,
                        'current_stock': None,
                        'sales_last_2_months': None,
                        'product_type': None,
                        'production_needed': None,
                        'production_priority': None,
                        'product_matched': False,
                        'has_product_data': False,
                        'is_in_tochka': False,
                        'needs_registration': True,
                    }
                
                merged_data.append(merged_item)
            
            # Сортируем: сначала требующие регистрации, потом по количеству заказов
            merged_data.sort(key=lambda x: (not x['needs_registration'], -x['orders']))
            
            coverage_rate = round((found_products / len(deduplicated_data)) * 100, 1) if deduplicated_data else 0
            
            analysis_result = {
                'message': f'Анализ производства завершен',
                'total_articles': len(deduplicated_data),
                'found_products': found_products,
                'coverage_rate': coverage_rate,
                'merged_data': merged_data
            }
            
        except Exception as e:
            return Response({
                'error': f'Ошибка при анализе производства: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Этап 3: Автоматическое формирование списка к производству
        try:
            # Получаем товары на производство из МойСклад
            products_for_production = Product.objects.filter(
                production_needed__gt=0
            ).order_by('-production_priority')
            
            # Создаем множество артикулов из Excel (товары Точки)
            tochka_articles = {item['article'] for item in deduplicated_data}
            
            filtered_production = []
            for product in products_for_production:
                normalized_article = normalize_article(product.article)
                
                if normalized_article in tochka_articles:
                    # Находим соответствующую запись из Excel для orders_in_tochka
                    excel_item = next((item for item in deduplicated_data if item['article'] == normalized_article), None)
                    orders_in_tochka = excel_item['orders'] if excel_item else 0
                    
                    # Рассчитываем резерв с новым алгоритмом
                    reserved_stock = float(getattr(product, 'reserved_stock', 0))
                    current_stock = float(product.current_stock)
                    
                    reserve_calculator = ReserveCalculatorService()
                    reserve_calc = reserve_calculator.calculate_reserve_display(
                        reserved_stock=reserved_stock,
                        current_stock=current_stock
                    )
                    reserve_ui = reserve_calculator.format_reserve_for_display(reserve_calc)
                    
                    item = {
                        'article': product.article,
                        'product_name': product.name,
                        'production_needed': float(product.production_needed),
                        'production_priority': product.production_priority,
                        'current_stock': current_stock,
                        'sales_last_2_months': float(product.sales_last_2_months),
                        'product_type': product.product_type,
                        'reserved_stock': reserved_stock,
                        'orders_in_tochka': orders_in_tochka,
                        'is_in_tochka': True,
                        'needs_registration': False,
                        
                        # Новые поля расчета резерва
                        'calculated_reserve': float(reserve_calc['calculated_reserve']),
                        'reserve_color': reserve_calc['color_indicator'],
                        'reserve_display_text': reserve_ui['display_text'],
                        'reserve_tooltip': reserve_ui['tooltip_text'],
                        'reserve_needs_attention': reserve_ui['needs_attention'],
                        
                        # Дополнительные поля для UI
                        'has_reserve': reserved_stock > 0,
                        'reserve_amount': reserved_stock,
                        'reserve_minus_stock': float(reserve_calc['calculated_reserve'])  # Обратная совместимость
                    }
                    filtered_production.append(item)
            
            production_result = {
                'message': f'Список к производству сформирован',
                'total_products': len(filtered_production),
                'products_in_tochka': len(filtered_production),
                'products_need_registration': 0,
                'filtered_production': filtered_production
            }
            
        except Exception as e:
            return Response({
                'error': f'Ошибка при формировании списка производства: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Финальный ответ с результатами всех этапов
        end_time = time.time()
        processing_time = round(end_time - start_time, 2)
        
        return Response({
            'success': True,
            'processing_time_seconds': processing_time,
            'upload_result': upload_result,
            'analysis_result': analysis_result,
            'production_result': production_result,
            'summary': {
                'excel_file_processed': True,
                'analysis_completed': True,
                'production_list_ready': True,
                'total_excel_records': len(deduplicated_data),
                'products_found_in_db': found_products,
                'coverage_percentage': coverage_rate,
                'production_items_count': len(filtered_production)
            }
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        end_time = time.time()
        processing_time = round(end_time - start_time, 2)
        
        return Response({
            'error': f'Ошибка при автоматической обработке: {str(e)}',
            'processing_time_seconds': processing_time
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
